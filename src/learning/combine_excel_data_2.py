"""
Скрипт для создания аугментированного датасета из Excel-файлов с добавлением случайных технических характеристик.

Скрипт считывает данные из всех Excel-файлов в каталоге data/, объединяет их,
случайно выбирает заданное количество строк, добавляет случайные технические характеристики
и создает новый Excel-файл с тремя листами: INPUT, EXTENDED и SIMPLIFIED.
"""

import os
import random
from pathlib import Path
import pandas as pd
from openpyxl import Workbook


def generate_random_specs():
    """
    Генерирует случайные технические характеристики или None.
    
    Returns:
        Кортеж (тип_спецификации, текст_спецификации) или (None, None), где:
        - тип_спецификации: 'txt', 'lst' или None
        - текст_спецификации: строка со спецификацией или None
        
        С вероятностью ~50% возвращает None (не добавлять спецификацию)
    """
    # Случайно решаем, добавлять ли спецификацию вообще (примерно 50% шанс)
    if random.random() < 0.5:
        return (None, None)
    
    # Список возможных текстовых спецификаций
    lst = [
        "зенковка 90 градусов",
        "для торцевых канавок",
        "для ширины паза 0,5 мм"
    ]
    
    # Случайно выбираем: либо генерируем txt, либо берем из lst
    if random.choice([True, False]):
        # Генерируем txt с случайными значениями от 15 до 125
        x = random.randint(15, 125)
        y = random.randint(15, 125)
        txt = f"Lобщ.={x} мм  Lраб={y}  мм"
        return ('txt', txt)
    else:
        # Выбираем случайный элемент из списка
        selected = random.choice(lst)
        return ('lst', selected)


def read_excel_files_pandas(data_dir):
    """
    Считывает данные из всех Excel-файлов в указанном каталоге, используя pandas.
    
    Args:
        data_dir: Путь к каталогу с входными Excel-файлами
        
    Returns:
        pandas.DataFrame с объединенными данными из всех файлов
    """
    data_path = Path(data_dir)
    
    # Получаем все Excel-файлы в каталоге
    excel_files = list(data_path.glob("*.xlsx"))
    
    if not excel_files:
        print(f"Предупреждение: В каталоге {data_dir} не найдено Excel-файлов")
        return pd.DataFrame()
    
    print(f"Найдено {len(excel_files)} Excel-файлов")
    
    all_dataframes = []
    
    for excel_file in excel_files:
        print(f"Обработка файла: {excel_file.name}")
        
        try:
            # Читаем Excel-файл с помощью pandas
            df = pd.read_excel(excel_file)
            
            # Проверяем наличие необходимых столбцов
            required_columns = [
                'Производитель',
                'Код заказа',
                'Наименование',
                'Обозначение',
                'Количество',
                'Единица измерения'
            ]
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                print(f"  Предупреждение: В файле {excel_file.name} отсутствуют столбцы: {missing_columns}")
                continue
            
            # Отбираем только необходимые столбцы
            df = df[required_columns]
            
            # Удаляем полностью пустые строки
            df = df.dropna(how='all')
            
            print(f"  Считано строк: {len(df)}")
            all_dataframes.append(df)
            
        except Exception as e:
            print(f"  Ошибка при обработке файла {excel_file.name}: {e}")
            continue
    
    if not all_dataframes:
        return pd.DataFrame()
    
    # Объединяем все DataFrame в один
    combined_df = pd.concat(all_dataframes, ignore_index=True)
    
    return combined_df


def select_random_rows_pandas(df, num_rows=300):
    """
    Случайно выбирает заданное количество строк из DataFrame.
    
    Args:
        df: pandas.DataFrame с данными
        num_rows: Количество строк для выбора (минимум 300)
        
    Returns:
        pandas.DataFrame с выбранными строками
    """
    if num_rows < 300:
        num_rows = 300
    
    if len(df) <= num_rows:
        # Если строк меньше или равно требуемому количеству, возвращаем все
        print(f"Всего строк: {len(df)}, используем все")
        return df.reset_index(drop=True)
    else:
        # Случайно выбираем num_rows строк
        print(f"Всего строк: {len(df)}, случайно выбираем {num_rows}")
        return df.sample(n=num_rows, random_state=None).reset_index(drop=True)


def create_output_file_openpyxl(df, output_path):
    """
    Создает выходной Excel-файл с тремя листами: INPUT, EXTENDED, SIMPLIFIED,
    используя openpyxl для записи.
    
    Args:
        df: pandas.DataFrame с данными для записи
        output_path: Путь к выходному файлу
    """
    # Создаем новую рабочую книгу
    wb = Workbook()
    
    # Удаляем стандартный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Создаем три листа
    ws_input = wb.create_sheet('INPUT')
    ws_extended = wb.create_sheet('EXTENDED')
    ws_simplified = wb.create_sheet('SIMPLIFIED')
    
    # === Лист INPUT ===
    # Заголовки для листа INPUT
    input_headers = ['п/п', 'Код заказа', 'Наименование', 'Кол-во', 'Ед.']
    ws_input.append(input_headers)
    
    # === Лист EXTENDED ===
    # Заголовки для листа EXTENDED
    extended_headers = [
        'Обозначение',
        'Наименование',
        'Производитель',
        'Единица измерения',
        'Количество',
        'Техническое задание'
    ]
    ws_extended.append(extended_headers)
    
    # === Лист SIMPLIFIED ===
    # Заголовки для листа SIMPLIFIED
    simplified_headers = [
        'Наименование',
        'Единица измерения',
        'Количество',
        'Техническое задание'
    ]
    ws_simplified.append(simplified_headers)
    
    # Заполняем данные построчно
    for idx, row in df.iterrows():
        # Получаем значения из DataFrame
        proizvoditel = str(row['Производитель']) if pd.notna(row['Производитель']) else ''
        kod_zakaza = str(row['Код заказа']) if pd.notna(row['Код заказа']) else ''
        naimenovanie = str(row['Наименование']) if pd.notna(row['Наименование']) else ''
        oboznachenie = str(row['Обозначение']) if pd.notna(row['Обозначение']) else ''
        kolichestvo = str(row['Количество']) if pd.notna(row['Количество']) else ''
        edinitsa_izmereniya = str(row['Единица измерения']) if pd.notna(row['Единица измерения']) else ''
        
        # Формируем базовое объединенное наименование (без технических характеристик)
        if naimenovanie and oboznachenie:
            base_name = f"{naimenovanie} {oboznachenie}"
        elif naimenovanie:
            base_name = naimenovanie
        elif oboznachenie:
            base_name = oboznachenie
        else:
            base_name = ''
        
        # Генерируем случайные технические характеристики для текущей строки
        spec_type, spec_text = generate_random_specs()
        
        # Формируем полное наименование для листа INPUT (с добавлением спецификации)
        # Если spec_text is None, значит спецификацию не добавляем
        if spec_text and base_name:
            full_name_input = f"{base_name} {spec_text}"
        elif base_name:
            full_name_input = base_name
        elif spec_text:
            full_name_input = spec_text
        else:
            full_name_input = ''
        
        # === Заполнение листа INPUT ===
        input_row = [
            idx + 1,  # п/п - порядковый номер (начинается с 1)
            kod_zakaza,  # Код заказа
            full_name_input,  # Наименование (с добавленной спецификацией)
            kolichestvo,  # Кол-во
            edinitsa_izmereniya  # Ед.
        ]
        ws_input.append(input_row)
        
        # === Заполнение листа EXTENDED ===
        extended_row = [
            oboznachenie,  # Обозначение
            naimenovanie,  # Наименование (базовое, без спецификации)
            proizvoditel,  # Производитель
            edinitsa_izmereniya,  # Единица измерения
            kolichestvo,  # Количество
            spec_text if spec_text else ''  # Техническое задание (та же спецификация, что в INPUT, или пусто)
        ]
        ws_extended.append(extended_row)
        
        # === Заполнение листа SIMPLIFIED ===
        simplified_row = [
            base_name,  # Наименование (базовое, без спецификации)
            edinitsa_izmereniya,  # Единица измерения
            kolichestvo,  # Количество
            spec_text if spec_text else ''  # Техническое задание (та же спецификация, что в INPUT, или пусто)
        ]
        ws_simplified.append(simplified_row)
    
    # Сохраняем файл
    wb.save(output_path)
    print(f"\nВыходной файл сохранен: {output_path}")
    print(f"Количество строк данных: {len(df)}")


def main():
    """
    Главная функция скрипта.
    """
    # Определяем пути
    script_dir = Path(__file__).parent
    data_dir = script_dir / 'datasets' / 'data'
    output_dir = script_dir / 'datasets'
    output_file = output_dir / 'output.xlsx'
    
    print("=" * 70)
    print("Скрипт генерации аугментированного датасета с техническими характеристиками")
    print("=" * 70)
    print(f"Каталог с входными данными: {data_dir}")
    print(f"Выходной файл: {output_file}")
    print()
    
    # Проверяем существование каталога с данными
    if not data_dir.exists():
        print(f"ОШИБКА: Каталог {data_dir} не существует")
        return
    
    # Считываем все данные из Excel-файлов с помощью pandas
    print("Шаг 1: Чтение данных из Excel-файлов (pandas)...")
    combined_df = read_excel_files_pandas(data_dir)
    
    if combined_df.empty:
        print("ОШИБКА: Не удалось считать данные из файлов")
        return
    
    print(f"\nШаг 2: Выбор случайных строк...")
    # Выбираем случайные строки (минимум 300)
    num_rows = 500
    selected_df = select_random_rows_pandas(combined_df, num_rows)
    
    # Создаем выходной каталог, если он не существует
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\nШаг 3: Создание выходного файла с аугментацией (openpyxl)...")
    # Создаем выходной файл с помощью openpyxl
    create_output_file_openpyxl(selected_df, output_file)
    
    print("\n" + "=" * 70)
    print("Обработка завершена успешно!")
    print("=" * 70)
    print("\nОсобенности сгенерированного датасета:")
    print("- Только некоторые строки (случайным образом ~50%) содержат технические характеристики")
    print("- Технические характеристики одинаковы на всех трёх листах (INPUT, EXTENDED, SIMPLIFIED)")
    print("- Используется два типа характеристик:")
    print("  1. Размерные: 'Lобщ.=X мм  Lраб=Y мм' (X, Y от 15 до 125)")
    print("  2. Описательные: 'зенковка 90 градусов', 'для торцевых канавок', 'для ширины паза 0,5 мм'")
    print("- Строки без технических характеристик имеют пустое поле 'Техническое задание'")


if __name__ == '__main__':
    main()

