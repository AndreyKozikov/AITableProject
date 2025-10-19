"""
Скрипт для объединения данных из нескольких Excel-файлов и создания структурированного выходного файла.

Скрипт считывает данные из всех Excel-файлов в каталоге data/, объединяет их,
случайно выбирает заданное количество строк и создает новый Excel-файл с тремя листами:
INPUT, EXTENDED и SIMPLIFIED.
"""

import os
import random
from pathlib import Path
from openpyxl import Workbook, load_workbook


def read_excel_files(data_dir):
    """
    Считывает данные из всех Excel-файлов в указанном каталоге.
    
    Args:
        data_dir: Путь к каталогу с входными Excel-файлами
        
    Returns:
        Список словарей с данными из всех файлов
    """
    all_rows = []
    data_path = Path(data_dir)
    
    # Получаем все Excel-файлы в каталоге
    excel_files = list(data_path.glob("*.xlsx"))
    
    if not excel_files:
        print(f"Предупреждение: В каталоге {data_dir} не найдено Excel-файлов")
        return all_rows
    
    print(f"Найдено {len(excel_files)} Excel-файлов")
    
    for excel_file in excel_files:
        print(f"Обработка файла: {excel_file.name}")
        
        try:
            # Загружаем рабочую книгу
            wb = load_workbook(excel_file, data_only=True)
            # Берем первый лист
            ws = wb.active
            
            # Получаем заголовки из первой строки
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Определяем индексы нужных столбцов
            col_indices = {}
            required_columns = [
                'Производитель',
                'Код заказа',
                'Наименование',
                'Обозначение',
                'Количество',
                'Единица измерения'
            ]
            
            for idx, header in enumerate(headers):
                if header in required_columns:
                    col_indices[header] = idx
            
            # Проверяем, что все необходимые столбцы найдены
            missing_columns = [col for col in required_columns if col not in col_indices]
            if missing_columns:
                print(f"  Предупреждение: В файле {excel_file.name} отсутствуют столбцы: {missing_columns}")
                continue
            
            # Считываем данные построчно (начиная со второй строки)
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Пропускаем полностью пустые строки
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                row_data = {
                    'Производитель': row[col_indices['Производитель']] if row[col_indices['Производитель']] else '',
                    'Код заказа': row[col_indices['Код заказа']] if row[col_indices['Код заказа']] else '',
                    'Наименование': row[col_indices['Наименование']] if row[col_indices['Наименование']] else '',
                    'Обозначение': row[col_indices['Обозначение']] if row[col_indices['Обозначение']] else '',
                    'Количество': row[col_indices['Количество']] if row[col_indices['Количество']] else '',
                    'Единица измерения': row[col_indices['Единица измерения']] if row[col_indices['Единица измерения']] else ''
                }
                
                all_rows.append(row_data)
                row_count += 1
            
            print(f"  Считано строк: {row_count}")
            wb.close()
            
        except Exception as e:
            print(f"  Ошибка при обработке файла {excel_file.name}: {e}")
            continue
    
    return all_rows


def select_random_rows(all_rows, num_rows=500):
    """
    Случайно выбирает заданное количество строк из общего списка.
    
    Args:
        all_rows: Список всех строк данных
        num_rows: Количество строк для выбора (минимум 500)
        
    Returns:
        Список выбранных строк
    """
    if num_rows < 500:
        num_rows = 500
    
    if len(all_rows) <= num_rows:
        # Если строк меньше или равно требуемому количеству, возвращаем все
        print(f"Всего строк: {len(all_rows)}, используем все")
        return all_rows
    else:
        # Случайно выбираем num_rows строк
        print(f"Всего строк: {len(all_rows)}, случайно выбираем {num_rows}")
        return random.sample(all_rows, num_rows)


def create_output_file(rows, output_path):
    """
    Создает выходной Excel-файл с тремя листами: INPUT, EXTENDED, SIMPLIFIED.
    
    Args:
        rows: Список строк данных для записи
        output_path: Путь к выходному файлу
    """
    # Создаем новую рабочую книгу
    wb = Workbook()
    
    # Удаляем стандартный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Создаем три листа
    ws_input = wb.create_sheet('INPUT')
    ws_extended = wb.create_sheet('EXTENDED')
    ws_simplified = wb.create_sheet('SIMPLIFIED')
    
    # === Лист INPUT ===
    # Заголовки для листа INPUT
    input_headers = ['П', 'Артикул', 'Наименование', 'Кол-во', 'Ед.']
    ws_input.append(input_headers)
    
    # === Лист EXTENDED ===
    # Заголовки для листа EXTENDED
    extended_headers = [
        'Обозначение',
        'Наименование',
        'Производитель',
        'Единица измерения',
        'Количество',
        'Техническое задание'
    ]
    ws_extended.append(extended_headers)
    
    # === Лист SIMPLIFIED ===
    # Заголовки для листа SIMPLIFIED
    simplified_headers = [
        'Наименование',
        'Единица измерения',
        'Количество',
        'Техническое задание'
    ]
    ws_simplified.append(simplified_headers)
    
    # Заполняем данные построчно
    for idx, row_data in enumerate(rows, start=1):
        # Формируем объединенное наименование
        naimenovanie = str(row_data['Наименование']).strip()
        oboznachenie = str(row_data['Обозначение']).strip()
        
        # Объединяем Наименование + Обозначение
        if naimenovanie and oboznachenie:
            full_name = f"{naimenovanie} {oboznachenie}"
        elif naimenovanie:
            full_name = naimenovanie
        elif oboznachenie:
            full_name = oboznachenie
        else:
            full_name = ''
        
        # === Заполнение листа INPUT ===
        input_row = [
            idx,  # П - порядковый номер
            row_data['Код заказа'],  # Артикул
            full_name,  # Наименование (объединенное)
            row_data['Количество'],  # Кол-во
            row_data['Единица измерения']  # Ед.
        ]
        ws_input.append(input_row)
        
        # === Заполнение листа EXTENDED ===
        extended_row = [
            row_data['Обозначение'],  # Обозначение
            row_data['Наименование'],  # Наименование
            row_data['Производитель'],  # Производитель
            row_data['Единица измерения'],  # Единица измерения
            row_data['Количество'],  # Количество
            row_data['Код заказа']  # Техническое задание
        ]
        ws_extended.append(extended_row)
        
        # === Заполнение листа SIMPLIFIED ===
        simplified_row = [
            full_name,  # Наименование (объединенное)
            row_data['Единица измерения'],  # Единица измерения
            row_data['Количество'],  # Количество
            ''  # Техническое задание (пусто)
        ]
        ws_simplified.append(simplified_row)
    
    # Сохраняем файл
    wb.save(output_path)
    print(f"\nВыходной файл сохранен: {output_path}")
    print(f"Количество строк данных: {len(rows)}")


def main():
    """
    Главная функция скрипта.
    """
    # Определяем пути
    script_dir = Path(__file__).parent
    data_dir = script_dir / 'datasets' / 'data'
    output_dir = script_dir / 'datasets'
    output_file = output_dir / 'output.xlsx'
    
    print("=" * 60)
    print("Скрипт объединения данных из Excel-файлов")
    print("=" * 60)
    print(f"Каталог с входными данными: {data_dir}")
    print(f"Выходной файл: {output_file}")
    print()
    
    # Проверяем существование каталога с данными
    if not data_dir.exists():
        print(f"ОШИБКА: Каталог {data_dir} не существует")
        return
    
    # Считываем все данные из Excel-файлов
    print("Шаг 1: Чтение данных из Excel-файлов...")
    all_rows = read_excel_files(data_dir)
    
    if not all_rows:
        print("ОШИБКА: Не удалось считать данные из файлов")
        return
    
    print(f"\nШаг 2: Выбор случайных строк...")
    # Выбираем случайные строки (минимум 500)
    num_rows = 500
    selected_rows = select_random_rows(all_rows, num_rows)
    
    # Создаем выходной каталог, если он не существует
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\nШаг 3: Создание выходного файла...")
    # Создаем выходной файл
    create_output_file(selected_rows, output_file)
    
    print("\n" + "=" * 60)
    print("Обработка завершена успешно!")
    print("=" * 60)


if __name__ == '__main__':
    main()

