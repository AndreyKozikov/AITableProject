"""Модуль экспорта данных в Excel.

Этот модуль предоставляет функциональность для экспорта JSON данных в Excel файлы
с добавлением эталонных данных из CSV файлов модельной директории.
"""

import json
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.utils.config import MODEL_DIR, OUT_DIR_LEARNING_DATA, PARSING_DIR
from src.utils.logging_config import get_logger

# Получение настроенного логгера
logger = get_logger(__name__)


def read_json_to_dataframe(json_path: Path) -> pd.DataFrame:
    """Прочитать JSON файл и преобразовать в DataFrame.
    
    Args:
        json_path: Путь к JSON файлу.
        
    Returns:
        DataFrame с данными из JSON или пустой DataFrame при ошибке.
    """
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Конвертируем в DataFrame
        if isinstance(data, list):
            df = pd.DataFrame(data)
        elif isinstance(data, dict):
            df = pd.DataFrame([data])
        else:
            logger.warning(f"Неподдерживаемый формат данных в {json_path}")
            return pd.DataFrame()
        
        logger.debug(f"Прочитан JSON {json_path.name}, форма: {df.shape}")
        return df
        
    except Exception as e:
        logger.error(f"Ошибка чтения JSON файла {json_path}: {e}")
        return pd.DataFrame()


def adjust_column_widths(excel_path: Path) -> None:
    """Автоматически подогнать ширину столбцов по содержимому во всех листах Excel файла.
    
    Args:
        excel_path: Путь к Excel файлу для обработки.
    """
    try:
        # Загружаем Excel файл
        wb = load_workbook(excel_path)
        
        # Обрабатываем каждый лист
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Проходим по всем столбцам
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Находим максимальную длину содержимого в столбце
                for cell in column:
                    try:
                        if cell.value:
                            # Вычисляем длину с учетом переносов строк
                            cell_value = str(cell.value)
                            # Берем максимальную длину строки, если есть переносы
                            lines = cell_value.split('\n')
                            cell_length = max(len(line) for line in lines)
                            max_length = max(max_length, cell_length)
                    except:
                        pass
                
                # Устанавливаем ширину столбца с небольшим запасом
                # Минимум 10, максимум 100 символов
                adjusted_width = min(max(max_length + 2, 10), 100)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            logger.debug(f"  Ширина столбцов подстроена для листа: {sheet_name}")
        
        # Сохраняем изменения
        wb.save(excel_path)
        logger.debug(f"Excel файл сохранен с подстроенными столбцами: {excel_path.name}")
        
    except Exception as e:
        logger.warning(f"Не удалось подстроить ширину столбцов для {excel_path.name}: {e}")


def read_csv_files_from_model_dir() -> List[tuple]:
    """Прочитать все CSV файлы из MODEL_DIR.
    
    Returns:
        Список кортежей (имя_листа, DataFrame) для каждого CSV файла.
    """
    csv_files = []
    
    try:
        # Получаем все CSV файлы из MODEL_DIR
        csv_paths = sorted(MODEL_DIR.glob("*.csv"))
        
        if not csv_paths:
            logger.warning(f"CSV файлы не найдены в {MODEL_DIR}")
            return []
        
        logger.info(f"Найдено {len(csv_paths)} CSV файлов в {MODEL_DIR}")
        
        for csv_path in csv_paths:
            try:
                # Читаем CSV с разделителем точка с запятой
                df = pd.read_csv(csv_path, sep=',', dtype=str, encoding='utf-8')
                
                # Формируем имя листа (имя файла в верхнем регистре без расширения)
                sheet_name = csv_path.stem.upper()
                
                csv_files.append((sheet_name, df))
                logger.debug(f"CSV файл {csv_path.name} прочитан, лист: {sheet_name}, форма: {df.shape}")
                
            except Exception as e:
                logger.error(f"Ошибка чтения CSV файла {csv_path}: {e}")
                continue
        
        return csv_files
        
    except Exception as e:
        logger.error(f"Ошибка получения CSV файлов из {MODEL_DIR}: {e}")
        return []


def export_json_to_excel(json_path: Path, csv_sheets: List[tuple]) -> bool:
    """Экспортировать JSON файл в Excel с дополнительными листами из CSV.
    
    Args:
        json_path: Путь к исходному JSON файлу.
        csv_sheets: Список кортежей (имя_листа, DataFrame) из CSV файлов.
        
    Returns:
        True если экспорт успешен, False в противном случае.
    """
    try:
        # Читаем JSON в DataFrame
        df_input = read_json_to_dataframe(json_path)
        
        if df_input.empty:
            logger.warning(f"Пустые данные в {json_path.name}, пропускаем")
            return False
        
        # Формируем путь к выходному Excel файлу
        excel_path = OUT_DIR_LEARNING_DATA / f"{json_path.stem}.xlsx"
        
        logger.info(f"Создание Excel файла: {excel_path.name}")
        
        # Создаем Excel файл с использованием ExcelWriter
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Записываем лист INPUT с данными из JSON
            df_input.to_excel(writer, sheet_name='INPUT', index=False)
            logger.debug(f"  Записан лист INPUT ({df_input.shape[0]} строк, {df_input.shape[1]} столбцов)")
            
            # Записываем дополнительные листы из CSV файлов
            for sheet_name, df_csv in csv_sheets:
                df_csv.to_excel(writer, sheet_name=sheet_name, index=False)
                logger.debug(f"  Записан лист {sheet_name} ({df_csv.shape[0]} строк, {df_csv.shape[1]} столбцов)")
        
        # Подстраиваем ширину столбцов по содержимому
        adjust_column_widths(excel_path)
        
        logger.info(f"Excel файл успешно создан: {excel_path} ({1 + len(csv_sheets)} листов)")
        return True
        
    except Exception as e:
        logger.error(f"Ошибка экспорта {json_path.name} в Excel: {e}", exc_info=True)
        return False


def export_all_json_to_excel() -> int:
    """Экспортировать все JSON файлы из PARSING_DIR в Excel файлы.
    
    Для каждого JSON файла создается отдельный Excel файл с листом INPUT (данные из JSON)
    и дополнительными листами из CSV файлов MODEL_DIR.
    
    Returns:
        Количество успешно созданных Excel файлов.
    """
    try:
        logger.info("=" * 70)
        logger.info("Начало экспорта JSON файлов в Excel")
        logger.info("=" * 70)
        
        # Получаем все JSON файлы из PARSING_DIR
        json_files = sorted(PARSING_DIR.glob("*.json"))
        
        if not json_files:
            logger.warning(f"JSON файлы не найдены в {PARSING_DIR}")
            return 0
        
        logger.info(f"Найдено {len(json_files)} JSON файлов для экспорта")
        
        # Читаем все CSV файлы из MODEL_DIR один раз
        csv_sheets = read_csv_files_from_model_dir()
        logger.info(f"Загружено {len(csv_sheets)} CSV листов из MODEL_DIR")
        
        # Обрабатываем каждый JSON файл
        success_count = 0
        
        for json_path in json_files:
            logger.info(f"\nОбработка: {json_path.name}")
            
            if export_json_to_excel(json_path, csv_sheets):
                success_count += 1
        
        # Итоговая статистика
        logger.info("=" * 70)
        logger.info(f"Экспорт завершен: {success_count}/{len(json_files)} файлов успешно обработано")
        logger.info(f"Excel файлы сохранены в: {OUT_DIR_LEARNING_DATA}")
        logger.info("=" * 70)
        
        return success_count
        
    except Exception as e:
        logger.error(f"Критическая ошибка при экспорте JSON в Excel: {e}", exc_info=True)
        return 0


if __name__ == "__main__":
    # Запуск экспорта при прямом выполнении скрипта
    export_all_json_to_excel()

